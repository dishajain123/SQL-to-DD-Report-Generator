"""Architecture step 18: DD Excel Export.

Writes DD rows to .xlsx using the exact column schema observed in the
platform's own Derivations export (Entity Name, Column Name, Column Type,
Derivation Option, Display Derivation Expression, Effective Start Date,
Status, Data Type, Decision Table Json, Conditional Json).

Supports merging into an existing DD export: rows not touched by this run
are preserved untouched; rows this run regenerated (same Entity+Column+
Effective Start Date) are replaced, not duplicated. This matters because a
DD Excel is a living document — re-running the pipeline on one changed
procedure should not wipe out everything else the team has already
reviewed and accepted.

The existing DD export can be supplied either as a real .xlsx/.xlsm
workbook or as a .csv export of the exact same column schema (e.g. a CSV
saved out of Excel) — both are read into the same internal row shape before
merging, so which format the platform's own DD export happens to be saved
in does not matter as long as the column names/order match.
"""
from __future__ import annotations

import csv
import json
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.core import DDRow
from app.utils import db

COLUMNS = [
    "Entity Name",
    "Column Name",
    "Column Type",
    "Derivation Option",
    "Display Derivation Expression",
    "Effective Start Date",
    "Status",
    "Data Type",
    "Decision Table Json",
    "Conditional Json",
]

_COLUMN_KEYS = {
    "Entity Name": "entity_name",
    "Column Name": "column_name",
    "Column Type": "column_type",
    "Derivation Option": "derivation_option",
    "Display Derivation Expression": "display_derivation_expression",
    "Effective Start Date": "effective_start_date",
    "Status": "status",
    "Data Type": "data_type",
    "Decision Table Json": "decision_table_json",
    "Conditional Json": "conditional_json",
}

# File extensions treated as CSV-style exports of the Derivations schema,
# rather than real spreadsheet workbooks.
_CSV_EXTENSIONS = {".csv"}


def dd_row_to_dict(dd: DDRow) -> dict:
    return {
        "entity_name": dd.entity_name,
        "column_name": dd.column_name,
        "column_type": dd.column_type.value,
        "derivation_option": dd.derivation_option.value,
        "display_derivation_expression": dd.display_derivation_expression,
        "effective_start_date": dd.effective_start_date.strftime("%d-%m-%Y"),
        "status": dd.status.value,
        "data_type": dd.data_type,
        "decision_table_json": dd.decision_table_json or "",
        "conditional_json": dd.conditional_json or "",
    }


def _row_dict_to_dd_row(row: dict) -> DDRow:
    effective_start = row.get("effective_start_date")
    if isinstance(effective_start, date):
        effective_date = effective_start
    else:
        effective_date = date.fromisoformat(str(effective_start))

    validation_errors = row.get("validation_errors") or []
    if isinstance(validation_errors, str):
        try:
            validation_errors = json.loads(validation_errors)
        except json.JSONDecodeError:
            validation_errors = [validation_errors]

    return DDRow(
        entity_name=str(row.get("entity_name", "")),
        column_name=str(row.get("column_name", "")),
        column_type=row.get("column_type") if isinstance(row.get("column_type"), str) else str(row.get("column_type", "Physical")),
        derivation_option=row.get("derivation_option") if isinstance(row.get("derivation_option"), str) else str(row.get("derivation_option", "Formula Expression")),
        display_derivation_expression=str(row.get("expression") or row.get("display_derivation_expression") or ""),
        effective_start_date=effective_date,
        status=row.get("status") if isinstance(row.get("status"), str) else str(row.get("status", "PENDING_REVIEW")),
        data_type=str(row.get("data_type") or ""),
        decision_table_json=row.get("decision_table_json") or None,
        conditional_json=row.get("conditional_json") or None,
        source_chain_id=str(row.get("chain_id") or ""),
        source_object_ids=[],
        confidence=float(row.get("confidence") or 0.0),
        validation_errors=list(validation_errors),
    )


def _row_key(row: dict) -> tuple:
    return (row["entity_name"], row["column_name"], row["effective_start_date"])


def _read_existing_dd_xlsx(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    rows = []
    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(header, row_cells))
        rows.append({_COLUMN_KEYS[h]: row_dict.get(h, "") for h in COLUMNS if h in _COLUMN_KEYS})
    return rows


def _read_existing_dd_csv(path: Path) -> list[dict]:
    rows = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row_dict in reader:
            rows.append(
                {_COLUMN_KEYS[h]: (row_dict.get(h) or "") for h in COLUMNS if h in _COLUMN_KEYS}
            )
    return rows


def read_existing_dd_excel(path: str | Path) -> list[dict]:
    """Read an existing DD export back into plain dict rows, keyed the same
    way dd_row_to_dict produces them, so it can be merged against new rows.

    Accepts either a real .xlsx/.xlsm workbook or a .csv export of the same
    column schema (matched purely by file extension — the column
    names/order in the file are what actually define compatibility, not
    the container format). Returns [] if the file doesn't exist yet (first
    run)."""
    path = Path(path)
    if not path.exists():
        return []

    if path.suffix.lower() in _CSV_EXTENSIONS:
        return _read_existing_dd_csv(path)
    return _read_existing_dd_xlsx(path)


def merge_dd_rows(existing: list[dict], new_rows: list[DDRow]) -> list[dict]:
    """New rows always win over an existing row with the same
    (entity, column, effective_start_date) key. Existing rows with no
    matching new row are preserved unchanged."""
    new_dicts = [dd_row_to_dict(r) for r in new_rows]
    new_keys = {_row_key(r) for r in new_dicts}
    preserved = [r for r in existing if _row_key(r) not in new_keys]
    return preserved + new_dicts


def export_dd_rows(
    dd_rows: list[DDRow], output_path: str | Path, existing_dd_path: str | Path | None = None
) -> Path:
    if existing_dd_path is not None:
        existing = read_existing_dd_excel(existing_dd_path)
        merged = merge_dd_rows(existing, dd_rows)
    else:
        merged = [dd_row_to_dict(r) for r in dd_rows]

    wb = Workbook()
    ws = wb.active
    ws.title = "Derivations"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col_idx, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row_dict in enumerate(merged, start=2):
        for col_idx, header in enumerate(COLUMNS, start=1):
            key = _COLUMN_KEYS[header]
            ws.cell(row=row_idx, column=col_idx, value=row_dict.get(key, ""))

    for col_idx in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 26
    ws.freeze_panes = "A2"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def export_dd_rows_csv(
    dd_rows: list[DDRow], output_path: str | Path, existing_dd_path: str | Path | None = None
) -> Path:
    if existing_dd_path is not None:
        existing = read_existing_dd_excel(existing_dd_path)
        merged = merge_dd_rows(existing, dd_rows)
    else:
        merged = [dd_row_to_dict(r) for r in dd_rows]

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writeheader()
        for row in merged:
            writer.writerow({header: row.get(_COLUMN_KEYS[header], "") for header in COLUMNS})
    return output_path


def export_reviewed_dd_rows_for_job(
    job_id: str, output_path: str | Path, db_path: str | None = None
) -> Path:
    rows = db.get_dd_rows_for_job(job_id, db_path=db_path)
    dd_rows = [_row_dict_to_dd_row(dict(row)) for row in rows]
    return export_dd_rows(dd_rows, output_path)


def export_reviewed_dd_rows_for_job_csv(
    job_id: str, output_path: str | Path, db_path: str | None = None
) -> Path:
    rows = db.get_dd_rows_for_job(job_id, db_path=db_path)
    dd_rows = [_row_dict_to_dd_row(dict(row)) for row in rows]
    return export_dd_rows_csv(dd_rows, output_path)
