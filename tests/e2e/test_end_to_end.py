from dataclasses import replace

from app.models.core import DDStatus, Intent, JobPlan
from app.orchestration.pipeline import build_pipeline
from app.utils import db


def test_full_pipeline_generate_dd(
    dpd_calculation_sql, maxdpd_sql, npa_date_sql, mock_llm_client, function_reference,
    tmp_db_path, tmp_path, monkeypatch,
):
    monkeypatch.setattr(db, "settings", replace(db.settings, sqlite_db_path=tmp_db_path))
    db.init_db(tmp_db_path)

    job_plan = JobPlan(
        job_id="job-e2e-1", intent=Intent.GENERATE_DD, company="Acme Bank",
        platform="4X", include_dd_excel=True,
    )
    pipeline = build_pipeline(llm_client=mock_llm_client)

    result = pipeline.invoke(
        {
            "job_plan": job_plan,
            "uploaded_files": {
                "dpd.sql": dpd_calculation_sql,
                "maxdpd.sql": maxdpd_sql,
                "npa.sql": npa_date_sql,
            },
            "function_reference": function_reference,
            "entity_name_map": {"AccountCal_Stg": "FCT_NPA_PRODUCT"},
        }
    )

    # Structural + lineage stages ran and correctly found the 3-object chain
    assert len(result["objects"]) == 3
    assert len(result["chains"]) == 1
    assert len(result["chains"][0].object_ids) == 3
    assert all(info.smart_chunks for info in result["structural_infos"].values())

    # DD generation ran because intent required it
    assert len(result["dd_rows"]) > 0
    assert any(r.status == DDStatus.ACTIVE for r in result["dd_rows"])
    assert any(r.advisory_notes for r in result["dd_rows"])

    # Report was produced and contains a DD Conditions section
    assert result["report_path"]
    report_text = open(result["report_path"]).read()
    assert "DD Conditions" in report_text
    assert "DPD_Calculation" in report_text or "MaxDPD_ReferencePeriod_Calculation" in report_text

    # Excel export happened because include_dd_excel was True
    assert result.get("excel_path")
    import openpyxl
    wb = openpyxl.load_workbook(result["excel_path"])
    ws = wb.active
    assert ws.cell(1, 1).value == "Entity Name"
    assert ws.max_row > 1


def test_full_pipeline_skips_dd_generation_for_explain_intent(
    dpd_calculation_sql, mock_llm_client, function_reference, tmp_db_path, monkeypatch,
):
    monkeypatch.setattr(db, "settings", replace(db.settings, sqlite_db_path=tmp_db_path))
    db.init_db(tmp_db_path)

    job_plan = JobPlan(
        job_id="job-e2e-2", intent=Intent.EXPLAIN, company="Acme Bank", platform="4X",
    )
    pipeline = build_pipeline(llm_client=mock_llm_client)

    result = pipeline.invoke(
        {
            "job_plan": job_plan,
            "uploaded_files": {"dpd.sql": dpd_calculation_sql},
            "function_reference": function_reference,
            "entity_name_map": {},
        }
    )

    # DD Generation must NOT run for an Explain-only job plan
    assert result["dd_rows"] == []
    assert result["report_path"]


def test_full_pipeline_persists_dd_rows_for_review(
    dpd_calculation_sql, maxdpd_sql, npa_date_sql, broken_llm_client, function_reference,
    tmp_db_path, monkeypatch,
):
    monkeypatch.setattr(db, "settings", replace(db.settings, sqlite_db_path=tmp_db_path))
    db.init_db(tmp_db_path)

    job_plan = JobPlan(
        job_id="job-e2e-3", intent=Intent.GENERATE_DD, company="Acme Bank", platform="4X",
    )
    pipeline = build_pipeline(llm_client=broken_llm_client)
    result = pipeline.invoke(
        {
            "job_plan": job_plan,
            "uploaded_files": {"dpd.sql": dpd_calculation_sql, "maxdpd.sql": maxdpd_sql, "npa.sql": npa_date_sql},
            "function_reference": function_reference,
            "entity_name_map": {},
        }
    )

    assert all(r.status == DDStatus.PENDING_REVIEW for r in result["dd_rows"])

    from app.review import review_store
    pending = review_store.list_pending(tmp_db_path)
    assert len(pending) == len(result["dd_rows"])
