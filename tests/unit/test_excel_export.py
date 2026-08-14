from datetime import date

import openpyxl

from app.models.core import ColumnType, DDRow, DDStatus, DerivationOption
from app.report.excel_export import COLUMNS, export_dd_rows


def _sample_row(**overrides) -> DDRow:
    defaults = dict(
        entity_name="FCT_NPA_PRODUCT",
        column_name="DPD_Overdue",
        column_type=ColumnType.PHYSICAL,
        derivation_option=DerivationOption.FORMULA_EXPRESSION,
        display_derivation_expression='IF(ISNOTEMPTY("A"."X"))THEN(1)ELSE(0)',
        effective_start_date=date(2026, 1, 1),
        status=DDStatus.ACTIVE,
        data_type="number",
        source_chain_id="chain-1",
    )
    defaults.update(overrides)
    return DDRow(**defaults)


def test_export_writes_correct_headers(tmp_path):
    out = export_dd_rows([_sample_row()], tmp_path / "dd.xlsx")
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, len(COLUMNS) + 1)]
    assert headers == COLUMNS


def test_export_writes_row_data(tmp_path):
    row = _sample_row()
    out = export_dd_rows([row], tmp_path / "dd.xlsx")
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert ws.cell(2, 1).value == "FCT_NPA_PRODUCT"
    assert ws.cell(2, 2).value == "DPD_Overdue"
    assert ws.cell(2, 4).value == "Formula Expression"
    assert ws.cell(2, 5).value == row.display_derivation_expression


def test_export_freezes_header_row(tmp_path):
    out = export_dd_rows([_sample_row()], tmp_path / "dd.xlsx")
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert ws.freeze_panes == "A2"


def test_export_handles_decision_table_rows(tmp_path):
    row = _sample_row(
        derivation_option=DerivationOption.DECISION_TABLE,
        display_derivation_expression="",
        decision_table_json='{"buckets": [{"max_dpd": 0, "label": "Standard"}]}',
    )
    out = export_dd_rows([row], tmp_path / "dd.xlsx")
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert ws.cell(2, 9).value == row.decision_table_json


def test_merge_preserves_unrelated_existing_rows(tmp_path):
    from app.report.excel_export import read_existing_dd_excel

    # First run: two unrelated columns.
    row_a = _sample_row(column_name="DPD_Overdue")
    row_b = _sample_row(column_name="DPD_Renewal")
    first_path = tmp_path / "dd.xlsx"
    export_dd_rows([row_a, row_b], first_path)

    # Second run: only regenerates DPD_Overdue (e.g. one proc changed).
    updated_a = _sample_row(
        column_name="DPD_Overdue",
        display_derivation_expression='IF(ISNOTEMPTY("A"."X"))THEN(1)ELSE(0)',
    )
    second_path = tmp_path / "dd_v2.xlsx"
    export_dd_rows([updated_a], second_path, existing_dd_path=first_path)

    merged = read_existing_dd_excel(second_path)
    by_column = {r["column_name"]: r for r in merged}

    # DPD_Renewal was untouched by the second run -- must still be present.
    assert "DPD_Renewal" in by_column
    # DPD_Overdue must be the NEW expression, not duplicated.
    assert sum(1 for r in merged if r["column_name"] == "DPD_Overdue") == 1
    assert by_column["DPD_Overdue"]["display_derivation_expression"] == updated_a.display_derivation_expression


def test_merge_with_no_existing_file_behaves_like_fresh_export(tmp_path):
    row = _sample_row()
    out = export_dd_rows([row], tmp_path / "dd.xlsx", existing_dd_path=tmp_path / "does_not_exist.xlsx")
    wb = openpyxl.load_workbook(out)
    assert wb.active.max_row == 2  # header + one row


def test_read_existing_dd_excel_returns_empty_for_missing_file(tmp_path):
    from app.report.excel_export import read_existing_dd_excel

    assert read_existing_dd_excel(tmp_path / "nope.xlsx") == []
